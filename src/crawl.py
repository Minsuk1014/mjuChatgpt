# from bs4 import BeautifulSoup as bs
# from pathlib import Path
# from openpyxl import Workbook
# import time
# import os
# import re
# import requests as rq
# import json
# import math

# def get_headers(key: str) -> dict[str, dict[str, str]] | None:
#     """ Get Headers """
#     JSON_FILE: str = 'json/headers.json'
#     with open(JSON_FILE, 'r', encoding='UTF-8') as file:
#         headers = json.loads(file.read())

#     try:
#         return headers[key]
#     except:
#         raise EnvironmentError(f'Set the {key}')

# class Coupang():
#     @staticmethod
#     def get_product_code(url: str) -> str:
#         prod_code: str = url.split('products/')[-1].split('?')[0]
#         return prod_code

#     @staticmethod
#     def get_soup_object(resp: rq.Response) -> bs:
#         return bs(resp.text, 'html.parser')

#     def __init__(self) -> None:
#         self.__headers: dict[str, str] = get_headers(key='headers')
#         self.base_review_url: str = 'https://www.coupang.com/vp/product/reviews'
#         self.sd = SaveData()

#     def get_title(self, prod_code: str) -> str:
#         url = f'https://www.coupang.com/vp/products/{prod_code}'
#         resp = rq.get(url=url, headers=self.__headers)
#         soup = self.get_soup_object(resp=resp)
#         return soup.select_one('h1.prod-buy-header__title').text.strip()

#     def start(self, url: str) -> None:
#         self.sd.create_directory()
#         self.__headers['referer'] = url  # URL을 사용할 수 있도록 설정

#         prod_code: str = self.get_product_code(url=url)
#         self.title: str = self.get_title(prod_code=prod_code)

#         # 테스트를 위해서 리뷰 페이지를 한 페이지로 제한 / 실제로 할 땐 20page정도면 될 듯 100개 리뷰 정도.. 아니면 너무 오래걸리더라
#         review_pages = 1
        
#         # Set payload
#         payloads = [{
#             'productId': prod_code,
#             'page': page,
#             'size': 5, # 리뷰 5개씩 불러오는거고 
#             'sortBy': 'ORDER_SCORE_ASC',
#             'ratings': '',
#             'q': '',
#             'viRoleCode': 2,
#             'ratingSummary': True
#         } for page in range(1, review_pages + 1)]
        
#         with rq.Session() as session:
#             [self.fetch(session=session, payload=payload) for payload in payloads]

#     def fetch(self, session: rq.Session, payload: dict) -> None:
#         now_page: str = payload['page']
#         print(f"\n[INFO] Start crawling page {now_page} ...\n")
#         with session.get(url=self.base_review_url, headers=self.__headers, params=payload) as response:
#             html = response.text
#             soup = bs(html, 'html.parser')
            
#             # 상품명
#             title = soup.select_one('h1.prod-buy-header__title')
#             title = '-' if title is None or title.text == '' else title.text.strip()

#             # Article Boxes
#             articles = soup.select('article.sdp-review__article__list')
#             for article in articles:
#                 dict_data: dict[str, str | int] = {}
                
#                 # 리뷰 날짜
#                 review_date = article.select_one('div.sdp-review__article__list__info__product-info__reg-date')
#                 review_date = '-' if review_date is None or review_date.text == '' else review_date.text.strip()
                
#                 # 구매자 이름
#                 user_name = article.select_one('span.sdp-review__article__list__info__user__name')
#                 user_name = '-' if user_name is None or user_name.text == '' else user_name.text.strip()

#                 # 평점
#                 rating = article.select_one('div.sdp-review__article__list__info__product-info__star-orange')
#                 rating = 0 if rating is None else int(rating.attrs['data-rating'])

#                 # 구매자 상품명
#                 prod_name = article.select_one('div.sdp-review__article__list__info__product-info__name')
#                 prod_name = '-' if prod_name is None or prod_name.text == '' else prod_name.text.strip()

#                 # 헤드라인(타이틀)
#                 headline = article.select_one('div.sdp-review__article__list__headline')
#                 headline = '등록된 헤드라인이 없습니다' if headline is None or headline.text == '' else headline.text.strip()

#                 # 리뷰 내용
#                 review_content = article.select_one('div.sdp-review__article__list__review > div')
#                 review_content = '등록된 리뷰내용이 없습니다' if review_content is None else re.sub('[\n\t]', '', review_content.text.strip())

#                 # 맛 만족도
#                 answer = article.select_one('span.sdp-review__article__list__survey__row__answer')
#                 answer = '맛 평가 없음' if answer is None or answer.text == '' else answer.text.strip()
                    
#                 dict_data['title'] = self.title
#                 dict_data['prod_name'] = prod_name
#                 dict_data['review_date'] = review_date
#                 dict_data['user_name'] = user_name
#                 dict_data['rating'] = rating
#                 dict_data['headline'] = headline
#                 dict_data['review_content'] = review_content
#                 dict_data['answer'] = answer
#                 self.sd.save(datas=dict_data)
#                 print(dict_data, '\n')
#             time.sleep(1)

#     @staticmethod
#     def clear_console() -> None:
#         command: str = 'clear' if os.name not in ('nt', 'dos') else 'cls'
#         os.system(command)

#     def calculate_total_pages(self, review_counts: int) -> int:
#         reviews_per_page: int = 5
#         return int(math.ceil(review_counts / reviews_per_page))

# class SaveData():
#     def __init__(self) -> None:
#         self.wb: Workbook = Workbook()
#         self.ws = self.wb.active
#         self.ws.append(['이름', '작성일자', '평점', '리뷰 내용', '맛 만족도'])
#         self.row: int = 2
#         self.dir_name: str = 'Coupang-reviews'
#         self.create_directory()

#     def create_directory(self) -> None:
#         if not os.path.exists(self.dir_name):
#             os.makedirs(self.dir_name)

#     def save(self, datas: dict[str, str | int]) -> None:
#         file_name: str = os.path.join(self.dir_name, datas['title'] + '.xlsx')
#         self.ws[f"A{self.row}"] = datas['user_name']
#         self.ws[f"B{self.row}"] = datas['review_date']
#         self.ws[f"C{self.row}"] = datas['rating']
#         self.ws[f"D{self.row}"] = datas['review_content']
#         self.ws[f"E{self.row}"] = datas['answer']
#         self.row += 1
#         self.wb.save(filename=file_name)

#     def __del__(self) -> None:
#         self.wb.close()



from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
import time
import os
import re
import requests as rq
import json
import math

def get_headers(key: str) -> dict[str, dict[str, str]] | None:
    """ Get Headers """
    JSON_FILE: str = 'json/headers.json'
    with open(JSON_FILE, 'r', encoding='UTF-8') as file:
        headers = json.loads(file.read())

    try:
        return headers[key]
    except:
        raise EnvironmentError(f'Set the {key}')

class Coupang():
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split('products/')[-1].split('?')[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, 'html.parser')

    def __init__(self) -> None:
        self.__headers: dict[str, str] = get_headers(key='headers')
        self.base_review_url: str = 'https://www.coupang.com/vp/product/reviews'
        self.sd = SaveData()

    def get_title(self, prod_code: str) -> str:
        url = f'https://www.coupang.com/vp/products/{prod_code}'
        resp = rq.get(url=url, headers=self.__headers)
        soup = self.get_soup_object(resp=resp)
        return soup.select_one('h1.prod-buy-header__title').text.strip()

    def start(self, url: str) -> None:
        self.sd.create_directory()
        self.__headers['referer'] = url  # URL을 사용할 수 있도록 설정

        prod_code: str = self.get_product_code(url=url)
        self.title: str = self.get_title(prod_code=prod_code)

        # 테스트를 위해서 리뷰 페이지를 한 페이지로 제한 / 실제로 할 땐 20page정도면 될 듯 100개 리뷰 정도.. 아니면 너무 오래걸리더라
        review_pages = 1
        
        # Set payload
        payloads = [{
            'productId': prod_code,
            'page': page,
            'size': 5, # 리뷰 5개씩 불러오는거고 
            'sortBy': 'ORDER_SCORE_ASC',
            'ratings': '',
            'q': '',
            'viRoleCode': 2,
            'ratingSummary': True
        } for page in range(1, review_pages + 1)]
        
        with rq.Session() as session:
            [self.fetch(session=session, payload=payload) for payload in payloads]

    def fetch(self, session: rq.Session, payload: dict) -> None:
        now_page: str = payload['page']
        print(f"\n[INFO] Start crawling page {now_page} ...\n")
        with session.get(url=self.base_review_url, headers=self.__headers, params=payload) as response:
            html = response.text
            soup = bs(html, 'html.parser')
            
            # 상품명
            title = soup.select_one('h1.prod-buy-header__title')
            title = '-' if title is None or title.text == '' else title.text.strip()

            # Article Boxes
            articles = soup.select('article.sdp-review__article__list')
            for article in articles:
                dict_data: dict[str, str | int] = {}
                

                # 리뷰 내용
                review_content = article.select_one('div.sdp-review__article__list__review > div')
                review_content = '등록된 리뷰내용이 없습니다' if review_content is None else re.sub('[\n\t]', '', review_content.text.strip())
                    
                dict_data['title'] = self.title
                dict_data['review_content'] = review_content

                self.sd.save(datas=dict_data)
                print(dict_data, '\n')
            time.sleep(1)

    @staticmethod
    def clear_console() -> None:
        command: str = 'clear' if os.name not in ('nt', 'dos') else 'cls'
        os.system(command)

    def calculate_total_pages(self, review_counts: int) -> int:
        reviews_per_page: int = 5
        return int(math.ceil(review_counts / reviews_per_page))

class SaveData():
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append(['리뷰 내용'])
        self.row: int = 2
        self.dir_name: str = 'Coupang-reviews'
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)

    def save(self, datas: dict[str, str | int]) -> None:
        file_name: str = os.path.join(self.dir_name, datas['title'] + '.xlsx')
        # self.ws[f"A{self.row}"] = datas['user_name']
        # self.ws[f"B{self.row}"] = datas['review_date']
        # self.ws[f"C{self.row}"] = datas['rating']
        self.ws[f"D{self.row}"] = datas['review_content']
        # self.ws[f"E{self.row}"] = datas['answer']
        self.row += 1
        self.wb.save(filename=file_name)

    def __del__(self) -> None:
        self.wb.close()
